import pandas as pd
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import os


# --- Helper function: Excel serial date ---
def date_to_excel_serial(dt):
    if pd.isna(dt):
        return None
    base = datetime(1899, 12, 30).date()

    if isinstance(dt, datetime):
        dt = dt.date()
    elif isinstance(dt, pd.Timestamp):
        dt = dt.to_pydatetime().date()
    elif isinstance(dt, date):
        pass
    else:
        return None

    return (dt - base).days


# --- Step 1: Read input file ---
input_file = r"E:\dil_copies\Document-Ageing-Report-\data\export.xls"

if input_file.endswith(".xls"):
    df = pd.read_excel(input_file, engine="xlrd")
else:
    df = pd.read_excel(input_file, engine="openpyxl")

print("✅ Data loaded. Columns:", df.columns.tolist())

# Normalize column names: replace spaces & dots with underscores
# (keeps original names otherwise; we will detect Company/Comapany later)
df.columns = (
    df.columns
    .str.replace(r"[ .]", "_", regex=True)  # replace spaces & dots
    .str.replace(r"_+", "_", regex=True)    # merge multiple underscores
    .str.strip("_")                         # remove leading/trailing underscores
)

# Ensure Document_Date column exists (safeguard)
if "Document_Date" in df.columns:
    df["Document_Date"] = pd.to_datetime(df["Document_Date"], errors="coerce")
else:
    df["Document_Date"] = pd.NaT

# --- Step 2: Today’s date & serial ---
today = datetime.now().date()
today_str = today.strftime("%d.%m.%Y")
today_serial = date_to_excel_serial(today)

# Add ageing column
df["Doc_Serial"] = df["Document_Date"].apply(date_to_excel_serial)
df["Doc_Ageing"] = df["Doc_Serial"].apply(lambda x: (today_serial - x) if x is not None else None)

# --- Step 3: Create workbook & Summary sheet ---
wb = openpyxl.Workbook()
summary_ws = wb.active
summary_ws.title = "Summary"

# Title row (kept in B2 as in your original)
summary_ws.cell(row=2, column=2).value = f"Document Ageing Report as at {today_str}"
summary_ws.cell(row=2, column=2).font = Font(bold=True, size=14)

# Headers (row 4)
summary_headers = [
    "Company", "Account", "Document_currency",
    "Amount_in_doc_curr", "Local_Currency", "Amount_in_local_currency"
]
header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
header_font = Font(bold=True)
header_alignment = Alignment(horizontal="center", vertical="center")


def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        try:
            column = col[0].column_letter  # Get the column name (A, B, C...)
        except Exception:
            continue
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)  # Add some padding
        ws.column_dimensions[column].width = adjusted_width


thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write summary headers (row 4) with center+middle alignment
for col, header in enumerate(summary_headers, start=2):
    cell = summary_ws.cell(row=4, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.border = thin_border
    cell.font = header_font
    cell.alignment = header_alignment

# Make header row a bit taller for vertical centering
summary_ws.row_dimensions[4].height = 20

# --- Step 4: Build summary content ---
# detect company column name (handle both 'Comapany' and 'Company')
company_col = next((c for c in ["Comapany", "Company"] if c in df.columns), None)
if company_col is None:
    raise KeyError("Neither 'Comapany' nor 'Company' column found in input file.")

group = df.groupby([company_col, "Account", "Document_currency", "Local_Currency"] )
sums = group.agg({
    "Amount_in_doc_curr": "sum",
    "Amount_in_local_currency": "sum"
}).reset_index()

# Filter out zero local currency amounts
sums = sums[abs(sums["Amount_in_local_currency"]) > 1e-5]

# Write data to summary (rows start at 5)
for i, row in sums.iterrows():
    r = i + 5
    summary_ws.cell(row=r, column=2).value = row[company_col]
    summary_ws.cell(row=r, column=3).value = row["Account"]
    summary_ws.cell(row=r, column=4).value = row["Document_currency"]
    summary_ws.cell(row=r, column=5).value = row["Amount_in_doc_curr"]
    summary_ws.cell(row=r, column=6).value = row["Local_Currency"]
    summary_ws.cell(row=r, column=7).value = row["Amount_in_local_currency"]
    # Apply border to each cell and ensure white background (no fill)
    for col in range(2, 8):
        cell = summary_ws.cell(row=r, column=col)
        cell.border = thin_border
        # Force no fill so data rows remain white
        cell.fill = PatternFill(fill_type=None)

# Auto-adjust column widths after writing summary
auto_adjust_column_width(summary_ws)

# --- Step 5: Create per-account sheets ---
unique_accounts = df["Account"].unique()

for account in unique_accounts:
    account_df = df[df["Account"] == account].copy()
    ws = wb.create_sheet(title=str(account))

    # Headers (row 1) – include extra ageing col
    headers = [
        company_col, "Account", "Document_Date", "Document_Type", "Text",
        "Document_currency", "Amount_in_doc_curr",
        "Local_Currency", "Amount_in_local_currency", "Doc_Ageing"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Make header row taller for vertical centering
    ws.row_dimensions[1].height = 20

    # Write rows
    for idx, row in enumerate(account_df.itertuples(index=False), start=2):
        row_dict = row._asdict()

        ws.cell(row=idx, column=1).value = row_dict.get(company_col)
        ws.cell(row=idx, column=2).value = row_dict.get("Account")

        doc_date = row_dict.get("Document_Date")
        if pd.notna(doc_date):
            ws.cell(row=idx, column=3).value = pd.to_datetime(doc_date).strftime("%d.%m.%Y")
        else:
            ws.cell(row=idx, column=3).value = None

        ws.cell(row=idx, column=4).value = row_dict.get("Document_Type")
        ws.cell(row=idx, column=5).value = row_dict.get("Text")
        ws.cell(row=idx, column=6).value = row_dict.get("Document_currency")
        ws.cell(row=idx, column=7).value = row_dict.get("Amount_in_doc_curr")
        ws.cell(row=idx, column=8).value = row_dict.get("Local_Currency")
        ws.cell(row=idx, column=9).value = row_dict.get("Amount_in_local_currency")
        ws.cell(row=idx, column=10).value = row_dict.get("Doc_Ageing")

        # Apply thin border to data rows (optional)
        for col_idx in range(1, 11):
            ws.cell(row=idx, column=col_idx).border = thin_border
            ws.cell(row=idx, column=col_idx).fill = PatternFill(fill_type=None)

    # Add total row at the end
    total_row = len(account_df) + 2
    ws.cell(row=total_row, column=7).value = account_df["Amount_in_doc_curr"].sum()
    ws.cell(row=total_row, column=9).value = account_df["Amount_in_local_currency"].sum()

    # Auto-adjust column widths for this sheet after writing
    auto_adjust_column_width(ws)

# --- Step 6: Save output ---
wb.save("Final Report.xlsx")
print("✅ Final Report.xlsx generated successfully!")
