import pandas as pd
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import os


# --- Helper function: Excel serial date ---
def date_to_excel_serial(dt):
    if pd.isna(dt):
        return None
    base = datetime(1899, 12, 30).date()
    
    if isinstance(dt, datetime):
        dt = dt.date()
    elif isinstance(dt, pd.Timestamp):
        dt = dt.to_pydatetime().date()
    elif isinstance(dt, date):
        pass
    else:
        return None
    
    return (dt - base).days


# --- Step 1: Read input file ---
input_file = r"E:\\dil_copies\\Document-Ageing-Report-\\data\\export.xls"

if input_file.endswith(".xls"):
    df = pd.read_excel(input_file, engine="xlrd")
else:
    df = pd.read_excel(input_file, engine="openpyxl")

print("✅ Data loaded. Columns:", df.columns.tolist())

# Normalize column names: replace spaces & dots with underscores
df.columns = (
    df.columns
    .str.replace(r"[ .]", "_", regex=True)  # replace spaces & dots
    .str.replace(r"_+", "_", regex=True)    # merge multiple underscores
    .str.strip("_")                         # remove leading/trailing underscores
)

# Parse document date if available
if "Document_Date" in df.columns:
    df["Document_Date"] = pd.to_datetime(df["Document_Date"], errors="coerce")

# --- Step 2: Today’s date & serial ---
today = datetime.now().date()
today_str = today.strftime("%d.%m.%Y")
today_serial = date_to_excel_serial(today)

# Add ageing column
df["Doc_Serial"] = df["Document_Date"].apply(date_to_excel_serial)
df["Doc_Ageing"] = today_serial - df["Doc_Serial"]

# --- Step 3: Create workbook & Summary sheet ---
wb = openpyxl.Workbook()
summary_ws = wb.active
summary_ws.title = "Summary"

# Title row
summary_ws.cell(row=2, column=2).value = f"Document Ageing Report as at {today_str}"
summary_ws.cell(row=2, column=2).font = Font(bold=True, size=14)

# Headers (row 4)
summary_headers = [
    "Company", "Account", "Document_currency", 
    "Amount_in_doc_curr", "Local_Currency", "Amount_in_local_currency"
]
header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_alignment = Alignment(horizontal="center", vertical="center")
center_bottom_align = Alignment(horizontal="center", vertical="bottom")

for col, header in enumerate(summary_headers, start=2):
    cell = summary_ws.cell(row=4, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.alignment = header_alignment

# --- Step 4: Build summary content ---
group = df.groupby(["Comapany", "Account", "Document_currency", "Local_Currency"])
sums = group.agg({
    "Amount_in_doc_curr": "sum",
    "Amount_in_local_currency": "sum"
}).reset_index()

# Filter out zero local currency amounts
sums = sums[abs(sums["Amount_in_local_currency"]) > 1e-5]

# Write data to summary
for i, row in sums.iterrows():
    r = i + 5
    summary_ws.cell(row=r, column=2).value = row["Comapany"]
    summary_ws.cell(row=r, column=3).value = row["Account"]
    summary_ws.cell(row=r, column=4).value = row["Document_currency"]
    summary_ws.cell(row=r, column=5).value = row["Amount_in_doc_curr"]
    summary_ws.cell(row=r, column=6).value = row["Local_Currency"]
    summary_ws.cell(row=r, column=7).value = row["Amount_in_local_currency"]

    # Apply borders to each cell
    for col in range(2, 8):
        summary_ws.cell(row=r, column=col).border = thin_border

    # Center + bottom align only for specified columns
    summary_ws.cell(row=r, column=2).alignment = center_bottom_align  # Company
    summary_ws.cell(row=r, column=4).alignment = center_bottom_align  # Document_currency
    summary_ws.cell(row=r, column=6).alignment = center_bottom_align  # Local_Currency

# Auto adjust column widths after filling
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

auto_adjust_column_width(summary_ws)

# --- Step 5: Create per-account sheets ---
unique_accounts = df["Account"].unique()

for account in unique_accounts:
    account_df = df[df["Account"] == account].copy()
    ws = wb.create_sheet(title=str(account))

    headers = [
        "Company", "Account", "Document_Date", "Document_Type", "Text",
        "Document_currency", "Amount_in_doc_curr", 
        "Local_Currency", "Amount_in_local_currency", "Doc_Ageing"
    ]

    # Header row formatting (center-middle)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.alignment = header_alignment

    # Data rows
    for idx, row in enumerate(account_df.itertuples(index=False), start=2):
        row_dict = row._asdict()
        ws.cell(row=idx, column=1).value = row_dict.get("Comapany")
        ws.cell(row=idx, column=2).value = row_dict.get("Account")

        doc_date = row_dict.get("Document_Date")
        if pd.notna(doc_date):
            ws.cell(row=idx, column=3).value = pd.to_datetime(doc_date).strftime("%d.%m.%Y")
        else:
            ws.cell(row=idx, column=3).value = None
        ws.cell(row=idx, column=4).value = row_dict.get("Document_Type")
        ws.cell(row=idx, column=5).value = row_dict.get("Text")
        ws.cell(row=idx, column=6).value = row_dict.get("Document_currency")
        ws.cell(row=idx, column=7).value = row_dict.get("Amount_in_doc_curr")
        ws.cell(row=idx, column=8).value = row_dict.get("Local_Currency")
        ws.cell(row=idx, column=9).value = row_dict.get("Amount_in_local_currency")
        ws.cell(row=idx, column=10).value = row_dict.get("Doc_Ageing")

        # Alignment: only for Document_Type (4), Document_currency (6), Local_Currency (8), Doc_Ageing (10)
        ws.cell(row=idx, column=4).alignment = center_bottom_align
        ws.cell(row=idx, column=6).alignment = center_bottom_align
        ws.cell(row=idx, column=8).alignment = center_bottom_align
        ws.cell(row=idx, column=10).alignment = center_bottom_align

    # Totals row
    total_row = len(account_df) + 2
    ws.cell(row=total_row, column=7).value = account_df["Amount_in_doc_curr"].sum()
    ws.cell(row=total_row, column=9).value = account_df["Amount_in_local_currency"].sum()

    auto_adjust_column_width(ws)

# --- Step 6: Save output ---
wb.save("Final Report.xlsx")
print("✅ Final Report.xlsx generated successfully!")
