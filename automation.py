import pandas as pd
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font

# --- Helper function: Excel serial date ---
def date_to_excel_serial(dt):
    if pd.isna(dt):
        return None
    base = datetime(1899, 12, 30).date()
    
    if isinstance(dt, datetime):
        dt = dt.date()
    elif isinstance(dt, pd.Timestamp):
        dt = dt.to_pydatetime().date()
    elif isinstance(dt, date):
        pass
    else:
        return None
    
    return (dt - base).days


# --- Step 1: Read input file ---
input_file = r"E:\dil_copies\Document-Ageing-Report-\data\export.xls"
df = pd.read_excel(input_file, engine="xlrd")
#print(df.columns)


# Normalize column names: replace spaces & dots with underscores
#df.columns = df.columns.str.replace(r"[ .]", "_", regex=True)

df.columns = (
    df.columns
    .str.replace(r"[ .]", "_", regex=True)  # replace spaces & dots
    .str.replace(r"_+", "_", regex=True)    # merge multiple underscores
    .str.strip("_")                         # remove leading/trailing underscores
)

# Check normalized columns
print("Normalized columns:", df.columns.tolist())



# Parse document date
if "Document_Date" in df.columns:
    df["Document_Date"] = pd.to_datetime(df["Document_Date"], errors="coerce")

# --- Step 2: Today’s date & serial ---
today = datetime.now().date()
today_str = today.strftime("%d.%m.%Y")
today_serial = date_to_excel_serial(today)

# Add ageing column
df["Doc_Serial"] = df["Document_Date"].apply(date_to_excel_serial)
df["Doc_Ageing"] = today_serial - df["Doc_Serial"]

# --- Step 3: Create workbook & Summary sheet ---
wb = openpyxl.Workbook()
summary_ws = wb.active
summary_ws.title = "Summary"

# Title row
summary_ws.cell(row=2, column=2).value = f"Document Ageing Report as at {today_str}"
summary_ws.cell(row=2, column=2).font = Font(bold=True, size=14)

# summary_ws.cell(row=1, column=1).value = "Document Ageing Report"
# summary_ws.cell(row=2, column=2).value = today_str   # B2 = today’s date

# Headers (row 3)
summary_headers = [
    "Company", "Account", "Document_currency", 
    "Amount_in_doc_curr", "Local_Currency", "Amount_in_local_currency"
]
for col, header in enumerate(summary_headers, start=1):
    summary_ws.cell(row=3, column=col).value = header

# --- Step 4: Build summary content ---
group = df.groupby(["Comapany", "Account", "Document_currency", "Local_Currency"])
sums = group.agg({
    "Amount_in_doc_curr": "sum",
    "Amount_in_local_currency": "sum"
}).reset_index()

# Filter out zero local currency amounts
sums = sums[abs(sums["Amount_in_local_currency"]) > 1e-5]

# Write data to summary
for i, row in sums.iterrows():
    summary_ws.cell(row=i+4, column=1).value = row["Comapany"]
    summary_ws.cell(row=i+4, column=2).value = row["Account"]
    summary_ws.cell(row=i+4, column=3).value = row["Document_currency"]
    summary_ws.cell(row=i+4, column=4).value = row["Amount_in_doc_curr"]
    summary_ws.cell(row=i+4, column=5).value = row["Local_Currency"]
    summary_ws.cell(row=i+4, column=6).value = row["Amount_in_local_currency"]

# --- Step 5: Create per-account sheets ---
unique_accounts = df["Account"].unique()

for account in unique_accounts:
    account_df = df[df["Account"] == account].copy()
    ws = wb.create_sheet(title=str(account))
    
    # Headers (row 1) – include extra ageing col
    headers = [
        "Company", "Account", "Document_Date", "Document_Type", "Text",
        "Document_currency", "Amount_in_doc_curr", 
        "Local_Currency", "Amount_in_local_currency", "Doc_Ageing"
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    
    # Write rows
    for idx, row in enumerate(account_df.itertuples(index=False), start=2):
        row_dict = row._asdict()
        
        ws.cell(row=idx, column=1).value = row_dict.get("Comapany")
        ws.cell(row=idx, column=2).value = row_dict.get("Account")
        
        doc_date = row_dict.get("Document_Date")
        if pd.notna(doc_date):
            ws.cell(row=idx, column=3).value = pd.to_datetime(doc_date).strftime("%d.%m.%Y")
        else:
            ws.cell(row=idx, column=3).value = None
        
        ws.cell(row=idx, column=4).value = row_dict.get("Document_Type")
        ws.cell(row=idx, column=5).value = row_dict.get("Text")
        ws.cell(row=idx, column=6).value = row_dict.get("Document_currency")
        ws.cell(row=idx, column=7).value = row_dict.get("Amount_in_doc_curr")
        ws.cell(row=idx, column=8).value = row_dict.get("Local_Currency")
        ws.cell(row=idx, column=9).value = row_dict.get("Amount_in_local_currency")
        ws.cell(row=idx, column=10).value = row_dict.get("Doc_Ageing")
    
    # Add total row at the end
    total_row = len(account_df) + 2
    ws.cell(row=total_row, column=7).value = account_df["Amount_in_doc_curr"].sum()
    ws.cell(row=total_row, column=9).value = account_df["Amount_in_local_currency"].sum()

# --- Step 6: Save output ---
wb.save("Final Report.xlsx")
print("✅ Final Report.xlsx generated successfully!")
