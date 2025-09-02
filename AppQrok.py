import pandas as pd
from datetime import datetime, date
import openpyxl
from openpyxl.utils import get_column_letter

def date_to_excel_serial(dt):
    base = date(1900, 1, 1)
    return (dt - base).days + 1

# Read the export.xls file
# Assuming it has a sheet named 'Sheet1'
df = pd.read_excel('export.xls', sheet_name='Sheet1', engine='xlrd')  # Use xlrd for .xls files

# Parse dates with dayfirst=True since format is DD.MM.YYYY
df['Document Date'] = pd.to_datetime(df['Document Date'], dayfirst=True)
df['Entry Date'] = pd.to_datetime(df['Entry Date'], dayfirst=True)  # If needed, but not used

# Get today's date
today = datetime.now().date()
today_str = today.strftime('%d.%m.%Y')
today_serial = date_to_excel_serial(today)

# Add serial and ageing columns to df
df['Doc Serial'] = df['Document Date'].apply(lambda x: date_to_excel_serial(x.date()))
df['Doc Ageing'] = today_serial - df['Doc Serial']

# Create new workbook
wb = openpyxl.Workbook()

# Summary sheet
summary_ws = wb.active
summary_ws.title = 'Summary'
summary_ws.cell(row=1, column=1).value = f"Document Ageing Report as at {today_str}"

# Summary headers row 3
summary_headers = ['Comapany', 'Account', 'Document currency', 'Amount in doc. curr.', 'Local Currency', 'Amount in local currency']
for col, header in enumerate(summary_headers, start=1):
    summary_ws.cell(row=3, column=col).value = header

# Group by company and account for summary
group = df.groupby(['Comapany', 'Account', 'Document currency', 'Local Currency'])
sums = group.agg({
    'Amount in doc. curr.': 'sum',
    'Amount in local currency': 'sum'
}).reset_index()

# Filter non-zero (using abs > 1e-5 to handle floating point)
sums = sums[abs(sums['Amount in local currency']) > 1e-5]

# Write to summary starting row 4
for i, row in sums.iterrows():
    summary_ws.cell(row=i+4, column=1).value = row['Comapany']
    summary_ws.cell(row=i+4, column=2).value = row['Account']
    summary_ws.cell(row=i+4, column=3).value = row['Document currency']
    summary_ws.cell(row=i+4, column=4).value = row['Amount in doc. curr.']
    summary_ws.cell(row=i+4, column=5).value = row['Local Currency']
    summary_ws.cell(row=i+4, column=6).value = row['Amount in local currency']

# Now, for each unique account, create sheet
unique_accounts = df['Account'].unique()

for account in unique_accounts:
    account_df = df[df['Account'] == account].copy()
    
    # Create sheet
    ws = wb.create_sheet(title=str(account))
    
    # Headers row 1
    headers = ['Comapany', 'Account', 'Document Date', 'Document Type', 'Document currency', 'Amount in doc. curr.', 'Local Currency', 'Amount in local currency', 'Text', 'Doc Ageing']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    
    # Write rows
    for r, row in account_df.iterrows():
        ws.cell(row=r+2, column=1).value = row['Comapany']
        ws.cell(row=r+2, column=2).value = row['Account']
        ws.cell(row=r+2, column=3).value = row['Doc Serial']  # Serial number
        ws.cell(row=r+2, column=4).value = row['Document Type']
        ws.cell(row=r+2, column=5).value = row['Document currency']
        ws.cell(row=r+2, column=6).value = row['Amount in doc. curr.']
        ws.cell(row=r+2, column=7).value = row['Local Currency']
        ws.cell(row=r+2, column=8).value = row['Amount in local currency']
        ws.cell(row=r+2, column=9).value = row['Text']
        ws.cell(row=r+2, column=10).value = row['Doc Ageing']
    
    # Total row
    total_row = len(account_df) + 2
    ws.cell(row=total_row, column=1).value = None
    ws.cell(row=total_row, column=2).value = None
    ws.cell(row=total_row, column=3).value = None
    ws.cell(row=total_row, column=4).value = None
    ws.cell(row=total_row, column=5).value = None
    ws.cell(row=total_row, column=6).value = account_df['Amount in doc. curr.'].sum()
    ws.cell(row=total_row, column=7).value = None
    ws.cell(row=total_row, column=8).value = account_df['Amount in local currency'].sum()
    ws.cell(row=total_row, column=9).value = None
    ws.cell(row=total_row, column=10).value = None

# Save the new workbook
wb.save('Final Report.xlsx')